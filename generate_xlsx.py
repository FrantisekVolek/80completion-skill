#!/usr/bin/env python3
"""
Generate Excel report for 80% completion epic report.

Usage:
    python3 generate_xlsx.py <input_json> <output_dir>
"""

import glob
import json
import os
import sys

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# Args
# ============================================================
if len(sys.argv) < 3:
    print(f"Usage: {sys.argv[0]} <input_json> <output_dir>")
    sys.exit(1)

input_path = sys.argv[1]
output_dir = sys.argv[2]

with open(input_path) as f:
    data = json.load(f)

quarter = data["quarter"]
label = data["quarter_label"]
label_alt = data["quarter_label_alt"]
initiatives = data["initiatives"]

# Jira base URL for hyperlinks (stored by prefetch.py)
JIRA_BASE = data.get("jira_url", "").rstrip("/") + "/browse/"

# Filter out cancelled epics
labeled_epics = {k: v for k, v in data["labeled_epics"].items() if v.get("status") != "Cancelled"}
unlabeled_epics = {k: v for k, v in data["unlabeled_epics"].items() if v.get("status") != "Cancelled"}

# ============================================================
# Previous report (for Change column)
# ============================================================
prev_lab_counts = None
prev_unlab_counts = None
prev_lab_total = None
prev_unlab_total = None
prev_grand_total = None
prev_domain_data = {}  # domain -> {lab_counts, unlab_counts, lab_total, unlab_total, grand_total}

def _cat(status_name, status_category):
    if status_name == "Cancelled":
        return "Cancelled"
    if status_category == "Done":
        return "Done"
    if status_category == "In Progress":
        return "In Progress"
    return "To Do / Backlog"

def _count(epics):
    counts = {"Done": 0, "In Progress": 0, "To Do / Backlog": 0}
    for ep in epics:
        cat = _cat(ep.get("status", ""), ep.get("status_category", ""))
        if cat != "Cancelled":
            counts[cat] = counts.get(cat, 0) + 1
    return counts

# Find all _80completion_input.json files in sibling directories
parent_dir = os.path.dirname(output_dir)
if parent_dir:
    all_inputs = sorted(glob.glob(os.path.join(parent_dir, "*/_80completion_input.json")))
    # Exclude current report
    all_inputs = [p for p in all_inputs if os.path.abspath(p) != os.path.abspath(input_path)]
    if all_inputs:
        prev_path = all_inputs[-1]  # most recent previous
        try:
            with open(prev_path) as f:
                prev_data = json.load(f)
            prev_labeled = {k: v for k, v in prev_data["labeled_epics"].items() if v.get("status") != "Cancelled"}
            prev_unlabeled_all = {k: v for k, v in prev_data["unlabeled_epics"].items() if v.get("status") != "Cancelled"}
            prev_initiatives = prev_data["initiatives"]

            prev_lab_counts = _count(list(prev_labeled.values()))
            # Build per-initiative labeled/unlabeled mapping
            prev_init_labeled = {}
            prev_init_unlabeled = {}
            prev_labeled_mapped = set()
            for init in prev_initiatives:
                prev_init_labeled[init["key"]] = []
                prev_init_unlabeled[init["key"]] = []
                for ek in init["linked_epic_keys"]:
                    if ek in prev_labeled:
                        prev_init_labeled[init["key"]].append(prev_labeled[ek])
                        prev_labeled_mapped.add(ek)
                    elif ek in prev_unlabeled_all:
                        prev_init_unlabeled[init["key"]].append(prev_unlabeled_all[ek])
            # Deduplicate unlabeled linked epics
            prev_seen = set()
            prev_unique_ul = []
            for epics in prev_init_unlabeled.values():
                for ep in epics:
                    if ep["key"] not in prev_seen:
                        prev_seen.add(ep["key"])
                        prev_unique_ul.append(ep)
            prev_unlab_counts = _count(prev_unique_ul)
            prev_lab_total = sum(prev_lab_counts.values())
            prev_unlab_total = sum(prev_unlab_counts.values())
            prev_grand_total = prev_lab_total + prev_unlab_total

            # Per-domain previous data
            for init in prev_initiatives:
                domain = init.get("domain", "-")
                if domain not in prev_domain_data:
                    prev_domain_data[domain] = {"labeled": [], "unlabeled_seen": set(), "unlabeled": []}
                prev_domain_data[domain]["labeled"].extend(prev_init_labeled[init["key"]])
                for ep in prev_init_unlabeled[init["key"]]:
                    if ep["key"] not in prev_domain_data[domain]["unlabeled_seen"]:
                        prev_domain_data[domain]["unlabeled_seen"].add(ep["key"])
                        prev_domain_data[domain]["unlabeled"].append(ep)
            # Convert to counts
            for domain in prev_domain_data:
                d = prev_domain_data[domain]
                d["lab_counts"] = _count(d["labeled"])
                d["unlab_counts"] = _count(d["unlabeled"])
                d["lab_total"] = sum(d["lab_counts"].values())
                d["unlab_total"] = sum(d["unlab_counts"].values())
                d["grand_total"] = d["lab_total"] + d["unlab_total"]

            print(f"Previous report loaded: {prev_path}")
        except Exception as e:
            print(f"Warning: could not load previous report {prev_path}: {e}")

# ============================================================
# Helpers
# ============================================================

def categorize_status(status_name, status_category):
    if status_name == "Cancelled":
        return "Cancelled"
    if status_category == "Done":
        return "Done"
    if status_category == "In Progress":
        return "In Progress"
    return "To Do / Backlog"


def fmt_prio(v):
    if v is None:
        return "-"
    if isinstance(v, float) and v == int(v):
        return int(v)
    return v


# Sort initiatives by prio
initiatives.sort(key=lambda i: (i["prio"] is None, i["prio"] if i["prio"] is not None else 0))

# Map epics to initiatives
labeled_mapped = set()
init_labeled = {}
init_unlabeled = {}

for init in initiatives:
    init_labeled[init["key"]] = []
    init_unlabeled[init["key"]] = []
    for ek in init["linked_epic_keys"]:
        if ek in labeled_epics:
            init_labeled[init["key"]].append(labeled_epics[ek])
            labeled_mapped.add(ek)
        elif ek in unlabeled_epics:
            init_unlabeled[init["key"]].append(unlabeled_epics[ek])

unmapped = [labeled_epics[k] for k in labeled_epics if k not in labeled_mapped]

# Status breakdown
all_labeled = list(labeled_epics.values())
seen_ul = set()
unique_unlabeled = []
for epics in init_unlabeled.values():
    for ep in epics:
        if ep["key"] not in seen_ul:
            seen_ul.add(ep["key"])
            unique_unlabeled.append(ep)

def count_by_status(epics):
    counts = {"Done": 0, "In Progress": 0, "To Do / Backlog": 0}
    for ep in epics:
        cat = categorize_status(ep["status"], ep["status_category"])
        if cat != "Cancelled":
            counts[cat] += 1
    return counts

lab_counts = count_by_status(all_labeled)
unlab_counts = count_by_status(unique_unlabeled)
lab_total = sum(lab_counts.values())
unlab_total = sum(unlab_counts.values())
grand_total = lab_total + unlab_total

# ============================================================
# Styles
# ============================================================
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
amber_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
done_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
progress_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
todo_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
cancelled_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
bold_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
header_font = Font(bold=True, color="FFFFFF", size=10)
title_font = Font(bold=True, size=14)
section_font = Font(bold=True, size=12)
bold_font = Font(bold=True, size=10)
normal_font = Font(size=10)
link_font = Font(color="0563C1", underline="single", size=10)

change_pos_font = Font(color="006100", size=10)
change_neg_font = Font(color="9C0006", size=10)
change_zero_font = Font(color="808080", size=10)

STATUS_FILLS = {
    "Done": done_fill,
    "In Progress": progress_fill,
    "To Do / Backlog": todo_fill,
    "Cancelled": cancelled_fill,
}


def write_header_row(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=col_start + i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border


def write_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = normal_font
    cell.border = thin_border
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    return cell


def write_link_cell(ws, row, col, key):
    cell = ws.cell(row=row, column=col, value=key)
    if JIRA_BASE and not JIRA_BASE.endswith("//browse/"):
        cell.hyperlink = f"{JIRA_BASE}{key}"
    cell.font = link_font
    cell.border = thin_border
    cell.alignment = Alignment(vertical='top')
    return cell


def write_status_cell(ws, row, col, status):
    cell = write_cell(ws, row, col, status)
    if status == "Cancelled":
        cell.fill = cancelled_fill
    elif status in ("Done", "Closed"):
        cell.fill = done_fill
    elif "progress" in status.lower() or "deliver" in status.lower() or "roll" in status.lower():
        cell.fill = progress_fill
    return cell


def write_change_cell(ws, row, col, current_pct, prev_pct):
    """Write a change cell showing difference in percentage points."""
    if prev_pct is None:
        cell = write_cell(ws, row, col, "-")
        cell.font = change_zero_font
        cell.alignment = Alignment(horizontal='center', vertical='top')
        return cell
    diff = round(current_pct - prev_pct, 1)
    if diff > 0:
        cell = write_cell(ws, row, col, f"+{diff}pp")
        cell.font = change_pos_font
    elif diff < 0:
        cell = write_cell(ws, row, col, f"{diff}pp")
        cell.font = change_neg_font
    else:
        cell = write_cell(ws, row, col, "—")
        cell.font = change_zero_font
    cell.alignment = Alignment(horizontal='center', vertical='top')
    return cell


def write_section(ws, row, text):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = section_font
    cell.fill = section_fill
    for col in range(2, 10):
        ws.cell(row=row, column=col).fill = section_fill
    return row + 1


def write_count_change(ws, row, col, current, prev_val):
    """Write a count change cell (absolute difference, not percentage)."""
    if prev_val is None:
        c = write_cell(ws, row, col, "-")
        c.font = Font(bold=True, color="808080", size=10)
        c.alignment = Alignment(horizontal='center', vertical='top')
        return c
    diff = current - prev_val
    if diff > 0:
        c = write_cell(ws, row, col, f"+{diff}")
        c.font = Font(bold=True, color="006100", size=10)
    elif diff < 0:
        c = write_cell(ws, row, col, f"{diff}")
        c.font = Font(bold=True, color="9C0006", size=10)
    else:
        c = write_cell(ws, row, col, "—")
        c.font = Font(bold=True, color="808080", size=10)
    c.alignment = Alignment(horizontal='center', vertical='top')
    return c


def write_breakdown_table(ws, row, lab_c, unlab_c, lab_t, unlab_t, g_total,
                          prev_lab_c, prev_unlab_c, prev_lab_t, prev_unlab_t, prev_g_total):
    """Write a full status breakdown table with Change column. Returns next row."""
    def _pct(n, total):
        return round(n / total * 100, 1) if total else 0
    def _prev_pct(n, total):
        return round(n / total * 100, 1) if total and total > 0 else None

    has_prev = prev_g_total is not None and prev_g_total > 0

    for status in ["Done", "In Progress", "To Do / Backlog"]:
        write_cell(ws, row, 1, f"Labelled — {status}")
        write_cell(ws, row, 2, lab_c[status])
        write_cell(ws, row, 3, _pct(lab_c[status], g_total))
        write_change_cell(ws, row, 4, _pct(lab_c[status], g_total),
                          _prev_pct(prev_lab_c[status], prev_g_total) if has_prev else None)
        ws.cell(row=row, column=1).fill = STATUS_FILLS.get(status, todo_fill)
        row += 1

    for col, val in [(1, "Labelled total"), (2, lab_t), (3, _pct(lab_t, g_total))]:
        c = write_cell(ws, row, col, val)
        c.font = bold_font
        c.fill = bold_fill
    c = write_change_cell(ws, row, 4, _pct(lab_t, g_total),
                          _prev_pct(prev_lab_t, prev_g_total) if has_prev else None)
    c.font = Font(bold=True, color=c.font.color.rgb if c.font.color and c.font.color.rgb else "808080", size=10)
    c.fill = bold_fill
    row += 1

    for status in ["Done", "In Progress", "To Do / Backlog"]:
        write_cell(ws, row, 1, f"Unlabelled — {status}")
        write_cell(ws, row, 2, unlab_c[status])
        write_cell(ws, row, 3, _pct(unlab_c[status], g_total))
        write_change_cell(ws, row, 4, _pct(unlab_c[status], g_total),
                          _prev_pct(prev_unlab_c[status], prev_g_total) if has_prev else None)
        row += 1

    for col, val in [(1, "Unlabelled total"), (2, unlab_t), (3, _pct(unlab_t, g_total))]:
        c = write_cell(ws, row, col, val)
        c.font = bold_font
        c.fill = bold_fill
    c = write_change_cell(ws, row, 4, _pct(unlab_t, g_total),
                          _prev_pct(prev_unlab_t, prev_g_total) if has_prev else None)
    c.font = Font(bold=True, color=c.font.color.rgb if c.font.color and c.font.color.rgb else "808080", size=10)
    c.fill = bold_fill
    row += 1

    for col, val in [(1, "Total"), (2, g_total), (3, 100.0 if g_total else 0)]:
        c = write_cell(ws, row, col, val)
        c.font = bold_font
    write_count_change(ws, row, 4, g_total, prev_g_total if has_prev else None)
    row += 1

    return row


# ============================================================
# Sheet 1: Summary
# ============================================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Summary"

row = 1
ws.cell(row=row, column=1, value=f"{quarter} — 80% Completion Epic Report").font = title_font
row += 2

# Key stats
stats = [
    ("Labels searched", f"{label}, {label_alt}"),
    ("Committed initiatives", len(initiatives)),
    ("Labeled epics", lab_total),
    ("Unlabeled linked epics", unlab_total),
    ("Grand total epics", grand_total),
]
for label_text, val in stats:
    ws.cell(row=row, column=1, value=label_text).font = bold_font
    ws.cell(row=row, column=2, value=val).font = normal_font
    row += 1
row += 1

# Overall status breakdown table
row = write_section(ws, row, "Epic Status Breakdown")
write_header_row(ws, row, ["Category", "Count", "%", "Change"])
row += 1

row = write_breakdown_table(
    ws, row, lab_counts, unlab_counts, lab_total, unlab_total, grand_total,
    prev_lab_counts, prev_unlab_counts, prev_lab_total, prev_unlab_total, prev_grand_total
)
row += 1

# --- Per-Domain Breakdown ---
domains = sorted(set(init["domain"] for init in initiatives if init["domain"] != "-"))
if any(init["domain"] == "-" for init in initiatives):
    domains.append("-")

for domain in domains:
    domain_inits = [i for i in initiatives if i["domain"] == domain]
    domain_labeled = []
    domain_unlabeled_seen = set()
    domain_unlabeled = []
    for init in domain_inits:
        for ep in init_labeled[init["key"]]:
            domain_labeled.append(ep)
        for ep in init_unlabeled[init["key"]]:
            if ep["key"] not in domain_unlabeled_seen:
                domain_unlabeled_seen.add(ep["key"])
                domain_unlabeled.append(ep)

    d_lab_counts = count_by_status(domain_labeled)
    d_unlab_counts = count_by_status(domain_unlabeled)
    d_lab_total = sum(d_lab_counts.values())
    d_unlab_total = sum(d_unlab_counts.values())
    d_grand_total = d_lab_total + d_unlab_total

    pd = prev_domain_data.get(domain)
    has_prev = pd is not None and pd["grand_total"] > 0

    domain_label_text = domain if domain != "-" else "No Domain"
    row = write_section(ws, row, f"{domain_label_text} ({len(domain_inits)} initiatives, {d_grand_total} epics)")
    write_header_row(ws, row, ["Category", "Count", "%", "Change"])
    row += 1

    row = write_breakdown_table(
        ws, row, d_lab_counts, d_unlab_counts, d_lab_total, d_unlab_total, d_grand_total,
        pd["lab_counts"] if has_prev else None,
        pd["unlab_counts"] if has_prev else None,
        pd["lab_total"] if has_prev else None,
        pd["unlab_total"] if has_prev else None,
        pd["grand_total"] if has_prev else None,
    )
    row += 1

# Column widths
for col, w in {1: 22, 2: 10, 3: 8, 4: 10}.items():
    ws.column_dimensions[get_column_letter(col)].width = w
ws.freeze_panes = "A2"

# ============================================================
# Sheet: Committed Initiatives
# ============================================================
ws_init = wb.create_sheet("Committed Initiatives")
r = 1
ws_init.cell(row=r, column=1, value="Committed Initiatives").font = title_font
r += 2

write_header_row(ws_init, r, ["Prio", "Initiative", "Summary", "Link", "Owner", "Domain", "Status"])
r += 1

for init in initiatives:
    write_cell(ws_init, r, 1, fmt_prio(init["prio"]))
    write_cell(ws_init, r, 2, init["key"])
    write_cell(ws_init, r, 3, init["summary"])
    write_link_cell(ws_init, r, 4, init["key"])
    write_cell(ws_init, r, 5, init["owner"])
    write_cell(ws_init, r, 6, init["domain"])
    write_status_cell(ws_init, r, 7, init["status"])
    r += 1

for col, w in {1: 6, 2: 14, 3: 50, 4: 38, 5: 22, 6: 18, 7: 14}.items():
    ws_init.column_dimensions[get_column_letter(col)].width = w
ws_init.freeze_panes = "A4"

# ============================================================
# Sheet: Labeled Epics
# ============================================================
ws2 = wb.create_sheet("Labeled Epics")
row = 1
ws2.cell(row=row, column=1, value="Labeled Epics by Initiative").font = title_font
row += 2

write_header_row(ws2, row, ["Prio", "Initiative", "Domain", "Init. Summary", "Init. Link",
                             "Epic", "Epic Summary", "Epic Link", "Labels", "Epic Owner", "Status"])
row += 1

for init in initiatives:
    epics = init_labeled[init["key"]]
    if not epics:
        continue
    for ep in epics:
        write_cell(ws2, row, 1, fmt_prio(init["prio"]))
        write_cell(ws2, row, 2, init["key"])
        write_cell(ws2, row, 3, init["domain"])
        write_cell(ws2, row, 4, init["summary"])
        write_link_cell(ws2, row, 5, init["key"])
        write_cell(ws2, row, 6, ep["key"])
        write_cell(ws2, row, 7, ep["summary"])
        write_link_cell(ws2, row, 8, ep["key"])
        write_cell(ws2, row, 9, ", ".join(ep["labels"]))
        write_cell(ws2, row, 10, ep["epic_owner"])
        write_status_cell(ws2, row, 11, ep["status"])
        row += 1

for col, w in {1: 6, 2: 14, 3: 16, 4: 40, 5: 38, 6: 14, 7: 45, 8: 38, 9: 25, 10: 22, 11: 14}.items():
    ws2.column_dimensions[get_column_letter(col)].width = w
ws2.freeze_panes = "A4"

# ============================================================
# Sheet: Unlabeled Epics
# ============================================================
ws3 = wb.create_sheet("Unlabeled Epics")
row = 1
ws3.cell(row=row, column=1, value="Linked Epics Without Quarter Labels").font = title_font
row += 2

write_header_row(ws3, row, ["Initiative", "Domain", "Init. Summary", "Init. Link",
                             "Epic", "Epic Summary", "Epic Link", "Status"])
row += 1

for init in initiatives:
    epics = init_unlabeled[init["key"]]
    if not epics:
        continue
    for ep in epics:
        write_cell(ws3, row, 1, init["key"])
        write_cell(ws3, row, 2, init["domain"])
        write_cell(ws3, row, 3, init["summary"])
        write_link_cell(ws3, row, 4, init["key"])
        write_cell(ws3, row, 5, ep["key"])
        write_cell(ws3, row, 6, ep["summary"])
        write_link_cell(ws3, row, 7, ep["key"])
        write_status_cell(ws3, row, 8, ep["status"])
        row += 1

for col, w in {1: 14, 2: 16, 3: 40, 4: 38, 5: 14, 6: 45, 7: 38, 8: 14}.items():
    ws3.column_dimensions[get_column_letter(col)].width = w
ws3.freeze_panes = "A4"

# ============================================================
# Sheet: Unmapped & Issues
# ============================================================
ws4 = wb.create_sheet("Unmapped & Issues")
row = 1
ws4.cell(row=row, column=1, value="Epics with Unclear Initiative Mapping").font = title_font
row += 2

if unmapped:
    write_header_row(ws4, row, ["Epic", "Summary", "Link", "Labels", "Epic Owner", "Status"])
    row += 1
    for ep in unmapped:
        write_cell(ws4, row, 1, ep["key"])
        write_cell(ws4, row, 2, ep["summary"])
        write_link_cell(ws4, row, 3, ep["key"])
        write_cell(ws4, row, 4, ", ".join(ep["labels"]))
        write_cell(ws4, row, 5, ep["epic_owner"])
        write_status_cell(ws4, row, 6, ep["status"])
        row += 1
else:
    ws4.cell(row=row, column=1, value="All labeled epics are mapped to initiatives.").font = normal_font
    row += 1

row += 2

# Initiatives with no epics
ws4.cell(row=row, column=1, value="Initiatives with No Linked Epics").font = section_font
ws4.cell(row=row, column=1).fill = section_fill
for col in range(2, 7):
    ws4.cell(row=row, column=col).fill = section_fill
row += 1

no_epics = [i for i in initiatives if not i["linked_epic_keys"]]
if no_epics:
    write_header_row(ws4, row, ["Initiative", "Summary", "Link", "Owner", "Domain", "Status"])
    row += 1
    for init in no_epics:
        write_cell(ws4, row, 1, init["key"])
        write_cell(ws4, row, 2, init["summary"])
        write_link_cell(ws4, row, 3, init["key"])
        write_cell(ws4, row, 4, init["owner"])
        write_cell(ws4, row, 5, init["domain"])
        write_status_cell(ws4, row, 6, init["status"])
        row += 1
else:
    ws4.cell(row=row, column=1, value="All initiatives have at least one linked epic.").font = normal_font

for col, w in {1: 14, 2: 45, 3: 38, 4: 25, 5: 18, 6: 14}.items():
    ws4.column_dimensions[get_column_letter(col)].width = w

# ============================================================
# Save
# ============================================================
os.makedirs(output_dir, exist_ok=True)
q_name = quarter.split()[0]
from datetime import date as _date
xlsx_path = os.path.join(output_dir, f"{q_name}-Epic-Report_{_date.today().isoformat()}.xlsx")
wb.save(xlsx_path)
print(f"Excel saved: {xlsx_path}")
print(f"Sheets: {', '.join(wb.sheetnames)}")
